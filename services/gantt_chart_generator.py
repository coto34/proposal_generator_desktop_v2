import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime, timedelta
from typing import Dict, List, Any
import calendar

class GanttChartGenerator:
    """Generate professional Gantt charts in Excel with project management features"""
    
    def __init__(self, project_data: Dict, activities: List[Dict], budget_data: Dict):
        self.project = project_data
        self.activities = activities
        self.budget = budget_data
        self.colors = {
            'header': 'FF366092',
            'activity': 'FF4CAF50', 
            'milestone': 'FFFF5722',
            'critical': 'FFF44336',
            'completed': 'FF81C784',
            'delayed': 'FFFF8A65',
            'background': 'FFF5F5F5'
        }
    
    def generate_gantt_excel(self, output_path: str) -> bool:
        """Generate comprehensive Gantt chart in Excel"""
        try:
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create multiple sheets
            self._create_gantt_chart_sheet(wb)
            self._create_timeline_summary_sheet(wb)
            self._create_milestone_tracker_sheet(wb)
            self._create_resource_allocation_sheet(wb)
            self._create_budget_timeline_sheet(wb)
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error generating Gantt chart: {e}")
            return False
    
    def _create_gantt_chart_sheet(self, workbook):
        """Create main Gantt chart sheet"""
        ws = workbook.create_sheet("Cronograma Gantt")
        
        # Project information header
        self._add_project_header(ws)
        
        # Calculate project timeline
        start_date = datetime(2024, 1, 1)  # Would be dynamic from project data
        duration_months = self.project.get('duration_months', 24)
        
        # Create month headers
        header_row = 8
        month_headers = self._generate_month_headers(start_date, duration_months)
        
        # Activity information columns
        activity_columns = [
            ('B', 'Código WBS', 10),
            ('C', 'Actividad', 40), 
            ('D', 'Responsable', 15),
            ('E', 'Duración', 10),
            ('F', 'Inicio', 12),
            ('G', 'Fin', 12),
            ('H', 'Progreso', 10),
            ('I', 'Presupuesto', 15)
        ]
        
        # Set up column headers
        for col, header, width in activity_columns:
            ws[f'{col}{header_row}'] = header
            ws.column_dimensions[col].width = width
            cell = ws[f'{col}{header_row}']
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add month headers
        date_col_start = 10  # Column J
        for i, month_header in enumerate(month_headers):
            col_letter = self._get_column_letter(date_col_start + i)
            ws[f'{col_letter}{header_row}'] = month_header
            ws.column_dimensions[col_letter].width = 4
            
            cell = ws[f'{col_letter}{header_row}']
            cell.font = Font(bold=True, size=9, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        
        # Add activities and timeline bars
        current_row = header_row + 1
        
        for activity in self.activities:
            # Activity information
            ws[f'B{current_row}'] = activity.get('code', '')
            ws[f'C{current_row}'] = activity.get('name', '')
            ws[f'D{current_row}'] = activity.get('responsible', 'IEPADES')
            ws[f'E{current_row}'] = f"{activity.get('duration_months', 0)} meses"
            ws[f'F{current_row}'] = self._format_date(start_date, activity.get('start_month', 1))
            ws[f'G{current_row}'] = self._format_date(start_date, activity.get('end_month', 1))
            ws[f'H{current_row}'] = f"{activity.get('progress', 0)}%"
            
            # Calculate budget for this activity
            activity_budget = self._get_activity_budget(activity.get('code', ''))
            ws[f'I{current_row}'] = f"${activity_budget:,.0f}"
            
            # Add timeline bars
            self._add_timeline_bar(ws, current_row, activity, start_date, date_col_start, duration_months)
            
            # Style the row
            self._style_activity_row(ws, current_row, activity)
            
            current_row += 1
        
        # Add milestones
        milestones = self._generate_project_milestones()
        for milestone in milestones:
            ws[f'B{current_row}'] = milestone.get('code', '')
            ws[f'C{current_row}'] = f"🎯 {milestone.get('name', '')}"
            ws[f'D{current_row}'] = milestone.get('responsible', 'IEPADES')
            ws[f'E{current_row}'] = "Hito"
            ws[f'F{current_row}'] = milestone.get('date', '')
            ws[f'G{current_row}'] = milestone.get('date', '')
            ws[f'H{current_row}'] = "N/A"
            ws[f'I{current_row}'] = "$0"
            
            # Add milestone marker
            milestone_col = self._calculate_milestone_column(milestone.get('month', 1), date_col_start)
            milestone_cell = ws[f'{milestone_col}{current_row}']
            milestone_cell.value = "♦"
            milestone_cell.font = Font(size=16, color='FFFFFF')
            milestone_cell.fill = PatternFill(start_color=self.colors['milestone'], fill_type='solid')
            milestone_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        # Add summary statistics
        self._add_gantt_summary(ws, current_row + 2)
        
        # Apply final formatting
        self._apply_gantt_formatting(ws, header_row, current_row)
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to Excel column letter"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def _generate_month_headers(self, start_date: datetime, duration_months: int) -> List[str]:
        """Generate month headers for timeline"""
        headers = []
        current_date = start_date
        
        for i in range(duration_months):
            month_name = calendar.month_abbr[current_date.month]
            year = str(current_date.year)[2:]  # Two-digit year
            headers.append(f"{month_name}\n'{year}")
            
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        return headers
    
    def _format_date(self, start_date: datetime, month_offset: int) -> str:
        """Format date for display"""
        target_date = start_date + timedelta(days=(month_offset - 1) * 30)
        return target_date.strftime("%d/%m/%Y")
    
    def _get_activity_budget(self, activity_code: str) -> float:
        """Get budget allocated to specific activity"""
        activity_items = [
            item for item in self.budget.get('items', [])
            if item.get('activity_code') == activity_code
        ]
        return sum(item.get('total_cost', 0) for item in activity_items)
    
    def _add_timeline_bar(self, ws, row: int, activity: Dict, start_date: datetime, 
                         date_col_start: int, total_months: int):
        """Add timeline bar for activity"""
        start_month = activity.get('start_month', 1)
        end_month = activity.get('end_month', 1)
        progress = activity.get('progress', 0) / 100
        
        for month in range(1, total_months + 1):
            col_letter = self._get_column_letter(date_col_start + month - 1)
            cell = ws[f'{col_letter}{row}']
            
            if start_month <= month <= end_month:
                # Calculate progress for this month
                months_elapsed = month - start_month + 1
                total_duration = end_month - start_month + 1
                month_progress = min(1.0, max(0.0, (progress * total_duration - months_elapsed + 1)))
                
                if month_progress > 0.8:
                    # Completed
                    cell.fill = PatternFill(start_color=self.colors['completed'], fill_type='solid')
                    cell.value = "█"
                elif month_progress > 0.3:
                    # In progress
                    cell.fill = PatternFill(start_color=self.colors['activity'], fill_type='solid')
                    cell.value = "▓"
                elif month_progress > 0:
                    # Started
                    cell.fill = PatternFill(start_color=self.colors['activity'], fill_type='solid')
                    cell.value = "░"
                else:
                    # Planned
                    cell.fill = PatternFill(start_color=self.colors['activity'], fill_type='solid')
                    cell.value = "─"
                
                cell.font = Font(color='FFFFFF', size=8)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Mark critical path activities
                if activity.get('critical_path', False):
                    cell.fill = PatternFill(start_color=self.colors['critical'], fill_type='solid')
    
    def _style_activity_row(self, ws, row: int, activity: Dict):
        """Apply styling to activity row"""
        # Alternate row colors for better readability
        if row % 2 == 0:
            bg_color = 'FFF8F9FA'
        else:
            bg_color = 'FFFFFFFF'
        
        # Style information columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = ws[f'{col}{row}']
            cell.fill = PatternFill(start_color=bg_color, fill_type='solid')
            cell.border = Border(
                left=Side(style='thin', color='FFE0E0E0'),
                right=Side(style='thin', color='FFE0E0E0'),
                top=Side(style='thin', color='FFE0E0E0'),
                bottom=Side(style='thin', color='FFE0E0E0')
            )
            
            # Special formatting for specific columns
            if col == 'H':  # Progress column
                progress = activity.get('progress', 0)
                if progress >= 100:
                    cell.font = Font(color='FF2E7D32', bold=True)
                elif progress >= 50:
                    cell.font = Font(color='FFFF8F00')
                else:
                    cell.font = Font(color='FFD32F2F')
            
            if col == 'I':  # Budget column
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right')
    
    def _generate_project_milestones(self) -> List[Dict]:
        """Generate project milestones"""
        return [
            {
                'code': 'M1',
                'name': 'Inicio del Proyecto',
                'date': '01/01/2024',
                'month': 1,
                'responsible': 'IEPADES',
                'type': 'start'
            },
            {
                'code': 'M2',
                'name': 'Diagnóstico Completado',
                'date': '31/03/2024', 
                'month': 3,
                'responsible': 'Equipo Técnico',
                'type': 'deliverable'
            },
            {
                'code': 'M3',
                'name': 'Evaluación Medio Término',
                'date': '30/06/2024',
                'month': 12,
                'responsible': 'Evaluador Externo',
                'type': 'evaluation'
            },
            {
                'code': 'M4',
                'name': 'Finalización del Proyecto',
                'date': '31/12/2025',
                'month': 24,
                'responsible': 'IEPADES',
                'type': 'end'
            }
        ]
    
    def _calculate_milestone_column(self, month: int, date_col_start: int) -> str:
        """Calculate column letter for milestone"""
        return self._get_column_letter(date_col_start + month - 1)
    
    def _add_gantt_summary(self, ws, start_row: int):
        """Add project summary statistics"""
        summary_data = [
            ('Resumen del Proyecto', ''),
            ('Total de Actividades', len(self.activities)),
            ('Duración Total', f"{self.project.get('duration_months', 24)} meses"),
            ('Presupuesto Total', f"${sum(item.get('total_cost', 0) for item in self.budget.get('items', [])):,.0f}"),
            ('Beneficiarios Directos', self.project.get('beneficiaries_direct', 'N/A')),
            ('Progreso General', f"{self._calculate_overall_progress():.1f}%"),
            ('Actividades Críticas', len([a for a in self.activities if a.get('critical_path', False)])),
            ('Estado del Proyecto', self._get_project_status())
        ]
        
        for i, (label, value) in enumerate(summary_data):
            row = start_row + i
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            
            if i == 0:  # Header
                ws[f'B{row}'].font = Font(bold=True, size=12)
                ws[f'B{row}'].fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
                ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
    
    def _calculate_overall_progress(self) -> float:
        """Calculate overall project progress"""
        if not self.activities:
            return 0
        
        total_progress = sum(activity.get('progress', 0) for activity in self.activities)
        return total_progress / len(self.activities)
    
    def _get_project_status(self) -> str:
        """Determine project status"""
        progress = self._calculate_overall_progress()
        
        if progress < 25:
            return "🟡 Inicio"
        elif progress < 75:
            return "🟠 En Progreso"
        elif progress < 100:
            return "🔵 Finalizando"
        else:
            return "🟢 Completado"
    
    def _apply_gantt_formatting(self, ws, header_row: int, last_row: int):
        """Apply final formatting to Gantt chart"""
        # Freeze panes at activity information
        ws.freeze_panes = f'J{header_row + 1}'
        
        # Add conditional formatting for timeline
        timeline_range = f'J{header_row + 1}:{self._get_column_letter(60)}{last_row}'
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = self._get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_project_header(self, ws):
        """Add project information header"""
        # Merge cells for title
        ws.merge_cells('B1:I1')
        title_cell = ws['B1']
        title_cell.value = f"CRONOGRAMA DE PROYECTO: {self.project.get('title', 'Proyecto IEPADES')}"
        title_cell.font = Font(bold=True, size=16, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Project details
        details = [
            ('Donante:', self.project.get('donor', 'N/A')),
            ('Duración:', f"{self.project.get('duration_months', 24)} meses"),
            ('Ubicación:', f"{self.project.get('municipality', '')}, {self.project.get('department', '')}"),
            ('Responsable:', 'IEPADES - Instituto de Enseñanza para el Desarrollo Sostenible')
        ]
        
        for i, (label, value) in enumerate(details):
            row = 3 + i
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            ws[f'B{row}'].font = Font(bold=True)
    
    def _create_timeline_summary_sheet(self, workbook):
        """Create timeline summary sheet with key dates"""
        ws = workbook.create_sheet("Resumen Temporal")
        
        # Header
        ws['A1'] = "RESUMEN CRONOLÓGICO DEL PROYECTO"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Phase breakdown
        phases = [
            {
                'phase': 'Fase de Inicio',
                'duration': '3 meses',
                'activities': ['Conformación de equipo', 'Diagnóstico inicial', 'Plan de trabajo'],
                'budget': 25000,
                'key_deliverable': 'Plan de Implementación'
            },
            {
                'phase': 'Fase de Implementación',
                'duration': '18 meses',
                'activities': ['Capacitaciones', 'Asistencia técnica', 'Monitoreo'],
                'budget': 95000,
                'key_deliverable': 'Resultados de Proyecto'
            },
            {
                'phase': 'Fase de Cierre',
                'duration': '3 meses',
                'activities': ['Evaluación final', 'Sistematización', 'Transferencia'],
                'budget': 15000,
                'key_deliverable': 'Informe Final'
            }
        ]
        
        # Headers
        headers = ['Fase', 'Duración', 'Actividades Principales', 'Presupuesto', 'Entregable Clave']
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Phase data
        for i, phase in enumerate(phases):
            row = 4 + i
            ws.cell(row=row, column=1).value = phase['phase']
            ws.cell(row=row, column=2).value = phase['duration']
            ws.cell(row=row, column=3).value = ', '.join(phase['activities'])
            ws.cell(row=row, column=4).value = f"${phase['budget']:,}"
            ws.cell(row=row, column=5).value = phase['key_deliverable']
        
        # Critical path analysis
        ws.cell(row=8, column=1).value = "ANÁLISIS DE RUTA CRÍTICA"
        ws.cell(row=8, column=1).font = Font(bold=True, size=12)
        
        critical_activities = [a for a in self.activities if a.get('critical_path', False)]
        for i, activity in enumerate(critical_activities):
            row = 10 + i
            ws.cell(row=row, column=1).value = activity.get('code', '')
            ws.cell(row=row, column=2).value = activity.get('name', '')
            ws.cell(row=row, column=3).value = "CRÍTICA"
            ws.cell(row=row, column=3).font = Font(color='FFD32F2F', bold=True)
    
    def _create_milestone_tracker_sheet(self, workbook):
        """Create milestone tracking sheet"""
        ws = workbook.create_sheet("Hitos y Entregables")
        
        ws['A1'] = "SEGUIMIENTO DE HITOS Y ENTREGABLES"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = ['Hito', 'Fecha Planificada', 'Fecha Real', 'Estado', 'Responsable', 'Comentarios', 'Riesgo']
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Milestone data
        milestones = self._generate_detailed_milestones()
        for i, milestone in enumerate(milestones):
            row = 4 + i
            ws.cell(row=row, column=1).value = milestone['name']
            ws.cell(row=row, column=2).value = milestone['planned_date']
            ws.cell(row=row, column=3).value = milestone.get('actual_date', 'Pendiente')
            ws.cell(row=row, column=4).value = milestone['status']
            ws.cell(row=row, column=5).value = milestone['responsible']
            ws.cell(row=row, column=6).value = milestone.get('comments', '')
            ws.cell(row=row, column=7).value = milestone.get('risk_level', 'Bajo')
            
            # Color code status
            status_cell = ws.cell(row=row, column=4)
            if milestone['status'] == 'Completado':
                status_cell.fill = PatternFill(start_color='FFC8E6C9', fill_type='solid')
            elif milestone['status'] == 'En Progreso':
                status_cell.fill = PatternFill(start_color='FFFFF3E0', fill_type='solid')
            elif milestone['status'] == 'Atrasado':
                status_cell.fill = PatternFill(start_color='FFFFEBEE', fill_type='solid')
    
    def _generate_detailed_milestones(self) -> List[Dict]:
        """Generate detailed milestone list"""
        return [
            {
                'name': 'Firma de Convenio',
                'planned_date': '15/01/2024',
                'status': 'Completado',
                'responsible': 'Dirección IEPADES',
                'risk_level': 'Bajo'
            },
            {
                'name': 'Diagnóstico Participativo',
                'planned_date': '31/03/2024',
                'status': 'En Progreso',
                'responsible': 'Equipo Técnico',
                'comments': 'Avance del 75%',
                'risk_level': 'Medio'
            },
            {
                'name': 'Plan de Capacitación',
                'planned_date': '30/04/2024',
                'status': 'Planificado',
                'responsible': 'Especialista en Capacitación',
                'risk_level': 'Bajo'
            },
            {
                'name': 'Primera Evaluación',
                'planned_date': '30/06/2024',
                'status': 'Planificado',
                'responsible': 'Evaluador Externo',
                'risk_level': 'Medio'
            }
        ]
    
    def _create_resource_allocation_sheet(self, workbook):
        """Create resource allocation timeline"""
        ws = workbook.create_sheet("Asignación de Recursos")
        
        ws['A1'] = "CRONOGRAMA DE ASIGNACIÓN DE RECURSOS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:N1')
        
        # Resource types
        resources = [
            {'name': 'Coordinador de Proyecto', 'type': 'Personal', 'monthly_cost': 3150},
            {'name': 'Especialista Técnico', 'type': 'Personal', 'monthly_cost': 2500},
            {'name': 'Facilitadores Comunitarios', 'type': 'Personal', 'monthly_cost': 4000},
            {'name': 'Vehículos', 'type': 'Logística', 'monthly_cost': 800},
            {'name': 'Equipos de Computo', 'type': 'Tecnología', 'monthly_cost': 200}
        ]
        
        # Headers
        headers = ['Recurso', 'Tipo'] + [f'Mes {i+1}' for i in range(12)]
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=i+1)
            cell.value = header
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Resource allocation timeline
        for i, resource in enumerate(resources):
            row = 4 + i
            ws.cell(row=row, column=1).value = resource['name']
            ws.cell(row=row, column=2).value = resource['type']
            
            # Monthly allocation (simplified)
            for month in range(12):
                col = 3 + month
                ws.cell(row=row, column=col).value = f"${resource['monthly_cost']:,}"
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFE8F5E8', fill_type='solid')
    
    def _create_budget_timeline_sheet(self, workbook):
        """Create budget execution timeline"""
        ws = workbook.create_sheet("Flujo de Caja")
        
        ws['A1'] = "PROYECCIÓN DE FLUJO DE CAJA MENSUAL"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:N1')
        
        # Calculate monthly budget distribution
        monthly_budget = self._calculate_monthly_budget_distribution()
        
        # Headers
        headers = ['Categoría'] + [f'Mes {i+1}' for i in range(12)] + ['Total']
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Budget categories
        categories = ['Personal', 'Equipamiento', 'Viajes', 'Capacitación', 'Operativos']
        
        for i, category in enumerate(categories):
            row = 4 + i
            ws.cell(row=row, column=1).value = category
            
            total_category = 0
            for month in range(12):
                col = 2 + month
                amount = monthly_budget.get(category, {}).get(f'month_{month+1}', 0)
                ws.cell(row=row, column=col).value = amount
                ws.cell(row=row, column=col).number_format = '"$"#,##0'
                total_category += amount
            
            # Total column
            ws.cell(row=row, column=14).value = total_category
            ws.cell(row=row, column=14).number_format = '"$"#,##0'
            ws.cell(row=row, column=14).font = Font(bold=True)
        
        # Monthly totals row
        total_row = 4 + len(categories)
        ws.cell(row=total_row, column=1).value = "TOTAL MENSUAL"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        grand_total = 0
        for month in range(12):
            col = 2 + month
            monthly_total = sum(monthly_budget.get(cat, {}).get(f'month_{month+1}', 0) for cat in categories)
            ws.cell(row=total_row, column=col).value = monthly_total
            ws.cell(row=total_row, column=col).number_format = '"$"#,##0'
            ws.cell(row=total_row, column=col).font = Font(bold=True)
            ws.cell(row=total_row, column=col).fill = PatternFill(start_color='FFFFEB3B', fill_type='solid')
            grand_total += monthly_total
        
        # Grand total
        ws.cell(row=total_row, column=14).value = grand_total
        ws.cell(row=total_row, column=14).number_format = '"$"#,##0'
        ws.cell(row=total_row, column=14).font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=14).fill = PatternFill(start_color='FFFF5722', fill_type='solid')
    
    def _calculate_monthly_budget_distribution(self) -> Dict:
        """Calculate budget distribution by month and category"""
        # Simplified distribution - in practice this would be more sophisticated
        return {
            'Personal': {f'month_{i+1}': 8000 for i in range(12)},
            'Equipamiento': {f'month_{i+1}': 2000 if i < 3 else 500 for i in range(12)},
            'Viajes': {f'month_{i+1}': 1500 for i in range(12)},
            'Capacitación': {f'month_{i+1}': 3000 if i % 3 == 0 else 1000 for i in range(12)},
            'Operativos': {f'month_{i+1}': 800 for i in range(12)}
        }